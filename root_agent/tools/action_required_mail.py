import os
from typing import List, Dict, Any, Optional
from datetime import datetime, timezone
import openpyxl
import base64
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from .utils.tracker_config import get_tracker_path

from .utils.gmail_client import GmailAPIClient
from .utils.drive_templates import load_template_from_drive

TRACKER_FILENAME = "Onboarding EMail Tracker.xlsx"  # legacy reference; prefer get_tracker_path()

def _resolve_tracker_path() -> str:
    """Return tracker path via central config (preferred)."""
    return get_tracker_path()

def _load_html_template() -> str:
    """Load the Action Required email template from Google Drive or local file.
    Tries Drive first, then falls back to local templates directory.
    """
    template_filename = 'Action required.htm'
    
    # Try loading from Google Drive first
    try:
        template_content = load_template_from_drive(template_filename, use_cache=True)
        if template_content:
            print(f"[action_required_mail] Loaded template from Drive: {template_filename}")
            return template_content
    except Exception as e:
        print(f"[action_required_mail] Drive template load failed: {e}, falling back to local")
    
    # Fallback to local templates directory
    templates_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'templates')
    template_path = os.path.join(templates_dir, template_filename)
    
    if not os.path.exists(template_path):
        print(f"[action_required_mail] Template not found at {template_path}")
        return "<html><body><p>(Template missing)</p></body></html>"
    try:
        with open(template_path, 'r', encoding='utf-8', errors='replace') as f:
            print(f"[action_required_mail] Loaded template from local: {template_filename}")
            return f.read()
    except Exception as e:
        print(f"[action_required_mail] Failed to read template: {e}")
        return f"<html><body><p>(Template read error: {e})</p></body></html>"

def _update_action_required_timestamp(row_number: int) -> bool:
    try:
        path = _resolve_tracker_path()
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook.active
        
        timestamp = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')
        
        worksheet.cell(row=row_number, column=11, value=timestamp)
        
        workbook.save(path)
        workbook.close()
        return True
    except Exception as e:
        print(f"[action_required_mail] Failed to update timestamp for row {row_number}: {e}")
        return False

def _load_eligible_workers() -> List[Dict[str, Any]]:
    """Load workers who are eligible for action-required email.
    
    Criteria: Columns A through J (10 columns) are filled, and column K (11) is empty.
    Groups workers by Start Date (column N, index 13).
    
    Returns:
        List of dicts with 'start_date' and 'workers' (list of worker dicts)
    """
    path = _resolve_tracker_path()
    if not os.path.exists(path):
        raise FileNotFoundError(f"Tracker not found at path: {path}")
    
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook.active
    
    # Group workers by start date
    workers_by_date = {}
    
    for idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if idx == 1:  # Skip header
            continue
        
        # Check if row has data
        if not row or len(row) < 2:
            continue
            
        name = (row[0] or '').strip() if len(row) > 0 and row[0] else ''
        email = (row[1] or '').strip() if len(row) > 1 and row[1] else ''
        worker_id = (row[2] or '').strip() if len(row) > 2 and row[2] else ''  # Column C
        
        if not name or not email:
            continue
        
        # Check if columns A through J (indices 0-9) are filled
        columns_a_to_j_filled = all(
            (row[i] or '').strip() if len(row) > i and row[i] else False 
            for i in range(10)  # Columns A through J (indices 0-9)
        )
        
        # Check if column K (index 10) is empty (action required email not sent yet)
        action_required_sent = (row[10] or '').strip() if len(row) > 10 and row[10] and not isinstance(row[10], datetime) else ''
        
        # Get start date from column N (index 13)
        start_date_raw = row[13] if len(row) > 13 and row[13] else None
        
        # If eligible: columns A-J filled, K empty, and has a start date
        if columns_a_to_j_filled and not action_required_sent and start_date_raw:
            # Convert datetime to formatted string
            if isinstance(start_date_raw, datetime):
                start_date_str = start_date_raw.strftime('%d %B %Y')
            else:
                # Try to convert string to datetime first, then format
                try:
                    if isinstance(start_date_raw, str):
                        start_date_str = start_date_raw.strip()
                    else:
                        start_date_str = str(start_date_raw)
                except:
                    start_date_str = str(start_date_raw)
            
            if start_date_str not in workers_by_date:
                workers_by_date[start_date_str] = []
            
            workers_by_date[start_date_str].append({
                'name': name,
                'worker_id': worker_id,
                'email': email,
                'row': idx
            })
    
    workbook.close()
    
    # Convert to list format
    result = [
        {'start_date': date, 'workers': workers}
        for date, workers in workers_by_date.items()
    ]
    
    return result

def _generate_worker_rows_html(workers: List[Dict[str, str]]) -> str:
    rows_html = []
    for worker in workers:
        rows_html.append(f"""      <tr>
        <td>{worker['name']}</td>
        <td>{worker['worker_id']}</td>
      </tr>""")
    return '\n'.join(rows_html)

_GMAIL_CLIENT = None  # Cached instance

def _send_email_with_cc(client: GmailAPIClient, to: str, cc: List[str], subject: str, body: str, attachment_path: str = None, max_retries: int = 3) -> str:
    """Send email with CC recipients using Gmail API with retry logic.
    
    Parameters:
        client: GmailAPIClient instance
        to: Primary recipient email
        cc: List of CC recipient emails
        subject: Email subject
        body: HTML email body
        attachment_path: Optional path to attachment file
        max_retries: Maximum number of retry attempts
    
    Returns:
        Status string
    """
    import time
    
    for attempt in range(max_retries):
        try:
            # Create MIME message
            message = MIMEMultipart()
            message['to'] = to
            message['cc'] = ', '.join(cc)
            message['subject'] = subject
            message['from'] = 'me'
            
            # Add HTML body
            html_part = MIMEText(body, 'html')
            message.attach(html_part)
            
            # Add attachment if provided
            if attachment_path and os.path.exists(attachment_path):
                with open(attachment_path, 'rb') as f:
                    pdf_data = f.read()
                
                pdf_attachment = MIMEApplication(pdf_data, _subtype='pdf')
                pdf_attachment.add_header(
                    'Content-Disposition',
                    'attachment',
                    filename=os.path.basename(attachment_path)
                )
                message.attach(pdf_attachment)
            
            # Encode message
            raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode('utf-8')
            
            # Send via Gmail API
            result = client.service.users().messages().send(
                userId='me',
                body={'raw': raw_message}
            ).execute()
            
            return f"✅ Email sent successfully via Gmail API. Message ID: {result['id']}"
            
        except Exception as e:
            error_str = str(e).lower()
            retryable_errors = ['10053', 'connection', 'timeout', 'ssl', 'network', 'aborted', 'reset']
            is_retryable = any(err in error_str for err in retryable_errors)
            
            if is_retryable and attempt < max_retries - 1:
                wait_time = (attempt + 1) * 3  # 3, 6, 9 seconds
                print(f"[action_required_mail] Connection error on attempt {attempt + 1}/{max_retries}, retrying in {wait_time}s: {e}")
                time.sleep(wait_time)
                
                # Try to rebuild the Gmail service connection
                try:
                    client.service = None
                    client.authenticate()
                    print("[action_required_mail] Gmail service rebuilt successfully")
                except Exception as auth_err:
                    print(f"[action_required_mail] Failed to rebuild Gmail service: {auth_err}")
            else:
                return f"❌ Failed to send email with CC after {attempt + 1} attempts: {str(e)}"
    
    return f"❌ Failed to send email with CC after {max_retries} attempts"

def send_action_required_mail(start_date: str, workers: List[Dict[str, Any]], recipient_emails: Optional[List[str]] = None) -> dict:
    """Send action-required email for a group of workers with the same start date.
    All workers will be CC'd on the email.
    
    Parameters:
        start_date: The deadline date string (e.g., "3rd November")
        workers: List of worker dicts with 'name', 'worker_id', 'email', 'row'
        recipient_emails: Optional list of recipient emails. If None, sends to first worker with others CC'd.
    
    Returns:
        Dict with response status
    """
    global _GMAIL_CLIENT
    
    try:
        # Load template
        raw_template = _load_html_template()
        
        # Generate worker rows HTML
        worker_rows_html = _generate_worker_rows_html(workers)
        
        # Replace placeholders
        html = raw_template.replace('{Deadline_Date}', start_date)
        html = html.replace('{Worker_Rows}', worker_rows_html)
        
        subject = f"Action required by {start_date} | Google Onboarding to be completed"
        
        # Path to the attachment
        attachment_path = r"C:\Users\bhangupta\Downloads\Alphabet Onboarding\Alphabet Onboarding Guide.pdf"
        
        # Get all worker emails
        all_emails = [w['email'] for w in workers]
        
        # Determine primary recipient and CC list
        if recipient_emails is None:
            # Send to first worker, CC all others
            primary_recipient = all_emails[0] if all_emails else None
            cc_recipients = all_emails  # CC everyone including primary
        else:
            primary_recipient = recipient_emails[0] if recipient_emails else None
            cc_recipients = all_emails  # CC all workers
        
        if not primary_recipient:
            return {'response': 'Failed', 'error': 'No recipients found'}
        
        # Initialize Gmail client if needed
        if _GMAIL_CLIENT is None:
            try:
                _GMAIL_CLIENT = GmailAPIClient()
            except Exception as init_e:
                print(f"[action_required_mail] GmailAPIClient init failed: {init_e}")
                return {'response': 'Failed', 'error': f'Gmail client init failed: {init_e}'}
        
        try:
            gmail_status = _send_email_with_cc(
                client=_GMAIL_CLIENT,
                to=primary_recipient,
                cc=cc_recipients,
                subject=subject,
                body=html,
                attachment_path=attachment_path
            )
            print(f"[action_required_mail] Gmail API send status: {gmail_status}")
            
            if gmail_status and gmail_status.startswith("✅"):
                # Update timestamps for all workers
                for worker in workers:
                    _update_action_required_timestamp(worker['row'])
                
                return {
                    'response': 'Action Required Email Sent',
                    'start_date': start_date,
                    'total_workers': len(workers),
                    'primary_recipient': primary_recipient,
                    'cc_recipients': cc_recipients,
                    'status': gmail_status
                }
            else:
                return {
                    'response': 'Failed',
                    'error': gmail_status,
                    'start_date': start_date
                }
        except Exception as send_err:
            print(f"[action_required_mail] Failed to send email: {send_err}")
            return {'response': 'Failed', 'error': str(send_err)}
        
    except Exception as e:
        print(f"[action_required_mail] Unexpected error: {e}")
        return {'response': 'Failed', 'error': str(e)}

def batch_action_required_emails(dry_run: bool = False) -> Dict[str, Any]:
    """Send action-required emails to all eligible workers, grouped by start date.
    
    Parameters:
        dry_run: If True, only show what would be sent without actually sending
    
    Returns:
        Dict with batch results
    """
    try:
        worker_groups = _load_eligible_workers()
    except Exception as e:
        return {'response': 'Failed', 'error': str(e)}
    
    if not worker_groups:
        return {
            'response': 'No Eligible Workers',
            'message': 'No workers found that meet the criteria (columns A-J filled, K empty, with start date)'
        }
    
    if dry_run:
        return {
            'response': 'Dry Run',
            'groups': len(worker_groups),
            'worker_groups': [
                {
                    'start_date': group['start_date'],
                    'worker_count': len(group['workers']),
                    'workers': group['workers']
                }
                for group in worker_groups
            ]
        }
    
    results = []
    for group in worker_groups:
        result = send_action_required_mail(
            start_date=group['start_date'],
            workers=group['workers']
        )
        results.append(result)
    
    total_sent = sum(r.get('emails_sent', 0) for r in results)
    total_failed = sum(r.get('emails_failed', 0) for r in results)
    
    return {
        'response': 'Batch Complete',
        'groups_processed': len(worker_groups),
        'total_emails_sent': total_sent,
        'total_emails_failed': total_failed,
        'results': results
    }
