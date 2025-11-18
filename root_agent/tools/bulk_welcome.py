import os
from typing import List, Dict, Any, Optional
from .welcome_mail import welcome_mail
import openpyxl
from datetime import datetime, timezone
from .utils.tracker_config import get_tracker_path
from .utils.tracker_schema import (
    COLUMN_NAME,
    COLUMN_EMAIL_ID,
    COLUMN_WORKORDER_ID,
    COLUMN_LOCATION,
    COLUMN_TIMEZONE,
    COLUMN_CHROMEBOOK_SERIAL_NUMBER,
    COLUMN_EMAIL_STATUS,
    is_cell_filled,
)

TRACKER_FILENAME = "Onboarding EMail Tracker.xlsx"  # retained for backward compatibility; prefer get_tracker_path()

def _resolve_tracker_path() -> str:
    """Deprecated local resolver; delegates to tracker_config for path."""
    return get_tracker_path()

def _mark_welcome_sent(row_number: int) -> bool:
    """Write a status note into the EMAIL_STATUS column indicating welcome mail was sent.
    Uses COLUMN_EMAIL_STATUS from tracker_schema. Does NOT touch the Chromebook Serial Number column.
    """
    try:
        path = _resolve_tracker_path()
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook.active
        timestamp = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')
        worksheet.cell(row=row_number, column=COLUMN_EMAIL_STATUS, value=f"Welcome Sent {timestamp}")
        workbook.save(path)
        workbook.close()
        return True
    except Exception as e:
        print(f"[bulk_welcome] Failed to mark welcome sent for row {row_number}: {e}")
        return False

def _load_rows(limit: Optional[int] = None) -> List[Dict[str, str]]:
    """Select rows eligible for welcome email.

    Criteria per new schema:
    - Columns 1..5 (Name, Email ID, Workorder ID, Location, Timezone) must be filled.
    - Column 6 (Chromebook Serial Number) must be EMPTY (if filled we skip sending).
    Stops at first row where Name & Email are both empty (assumes end of data).
    Returns list of dicts with necessary candidate info.
    """
    path = _resolve_tracker_path()
    if not os.path.exists(path):
        raise FileNotFoundError(f"Tracker not found at path: {path}")

    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook.active

    required_indices = [
        COLUMN_NAME,
        COLUMN_EMAIL_ID,
        COLUMN_WORKORDER_ID,
        COLUMN_LOCATION,
        COLUMN_TIMEZONE,
    ]

    results: List[Dict[str, str]] = []
    for idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if idx == 1:  # Skip header
            continue
        # Convert 1-based to 0-based for accessing tuple 'row'
       
        # Fast termination condition
        name_val = row[COLUMN_NAME - 1] if len(row) >= COLUMN_NAME else None
        email_val = row[COLUMN_EMAIL_ID - 1] if len(row) >= COLUMN_EMAIL_ID else None
        if not is_cell_filled(name_val) and not is_cell_filled(email_val):
            break

        # Ensure all required filled
        all_filled = all(
            is_cell_filled(row[i - 1]) if len(row) >= i else False for i in required_indices
        )
        if not all_filled:
            continue

        # Chromebook Serial Number present? skip
        serial_val = row[COLUMN_CHROMEBOOK_SERIAL_NUMBER - 1] if len(row) >= COLUMN_CHROMEBOOK_SERIAL_NUMBER else None
        if is_cell_filled(serial_val):
            continue

        results.append({
            'name': str(row[COLUMN_NAME - 1]).strip(),
            'email': str(row[COLUMN_EMAIL_ID - 1]).strip(),
            'workorder_id': str(row[COLUMN_WORKORDER_ID - 1]).strip(),
            'location': str(row[COLUMN_LOCATION - 1]).strip(),
            'timezone': str(row[COLUMN_TIMEZONE - 1]).strip(),
            'row': idx,
        })
        if limit and len(results) >= limit:
            break
    workbook.close()
    return results

def tracker_welcome_emails(limit: Optional[int] = None, dry_run: bool = False) -> Dict[str, Any]:
    """Batch send welcome emails based on Excel tracker.
    """
    try:
        rows = _load_rows(limit=limit)
    except Exception as e:
        return {'response': 'Failed', 'error': str(e)}

    if dry_run:
        return {
            'response': 'Dry Run',
            'count': len(rows),
            'candidates': rows
        }

    successes: List[Dict[str, Any]] = []
    failures: List[Dict[str, Any]] = []
    for r in rows:
        try:
            res = welcome_mail(
                Candidate_Name=r['name'],
                Location=r.get('location', 'N/A'),  
                candidateEmailID=r['email']
            )
            if res.get('response') == 'Welcome Mail Sent':
                # Mark status instead of writing into serial number column
                timestamp_updated = _mark_welcome_sent(r['row'])
                successes.append({
                    'row': r['row'], 
                    'email': r['email'], 
                    'location': r.get('location'), 
                    'workorder_id': r.get('workorder_id'),
                    'timezone': r.get('timezone'),
                    'transport': res.get('transport'),
                    'timestamp_updated': timestamp_updated
                })
            else:
                failures.append({'row': r['row'], 'email': r['email'], 'location': r.get('location'), 'error': res.get('error'), 'gmail_status': res.get('gmail_status')})
        except Exception as send_err:
            failures.append({'row': r['row'], 'email': r['email'], 'location': r.get('location'), 'error': str(send_err)})

    return {
        'response': 'Batch Complete',
        'total': len(rows),
        'sent': len(successes),
        'failed': len(failures),
        'successes': successes,
        'failures': failures
    }
