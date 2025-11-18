import os
from datetime import datetime, timezone, timedelta
from openpyxl import load_workbook
from .utils.gmail_client import GmailAPIClient
from .utils.drive_templates import load_template_from_drive
from .utils.tracker_config import get_tracker_path

def _resolve_tracker_path():
    """Return tracker path via central config."""
    return get_tracker_path()

def _load_html_template():
    """Load the Compliance Documents Reminder HTML template from Google Drive or local file.
    Tries Drive first, then falls back to local templates directory.
    """
    template_filename = 'Compliance Documents Reminder.htm'
    
    # Try loading from Google Drive first
    try:
        template_content = load_template_from_drive(template_filename, use_cache=True)
        if template_content:
            print(f"[compliance_reminder] Loaded template from Drive: {template_filename}")
            return template_content
    except Exception as e:
        print(f"[compliance_reminder] Drive template load failed: {e}, falling back to local")
    
    # Fallback to local templates directory
    template_dir = os.path.join(
        os.path.dirname(os.path.dirname(__file__)), 
        'templates'
    )
    template_path = os.path.join(template_dir, template_filename)
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")
    
    with open(template_path, 'r', encoding='utf-8') as f:
        print(f"[compliance_reminder] Loaded template from local: {template_filename}")
        return f.read()

def _load_eligible_workers():
    """
    Load workers who need compliance document reminders:
    - Column K (11): Welcome email timestamp exists (email sent)
    - Column M (13): Partner Domain Account triggered = "No"
    - Column L (12): Compliance reminder timestamp is empty (reminder not sent)
    - Time check: At least 8 hours have passed since welcome email (Column K)
    
    Returns list of dicts with worker info.
    """
    tracker_path = _resolve_tracker_path()
    wb = load_workbook(tracker_path)
    ws = wb.active
    
    eligible_workers = []
    current_time = datetime.now(timezone.utc)
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        try:
            # Column K (index 10): Welcome email timestamp
            welcome_email_timestamp = row[10].value if len(row) > 10 else None
            
            # Column M (index 12): Partner Domain Account triggered
            partner_domain_triggered = row[12].value if len(row) > 12 else None
            
            # Column L (index 11): Compliance reminder sent timestamp
            compliance_reminder_sent = row[11].value if len(row) > 11 else None
            
            # Skip if welcome email not sent
            if not welcome_email_timestamp:
                continue
            
            # Skip if partner domain account already triggered
            if partner_domain_triggered and str(partner_domain_triggered).strip().upper() != "NO":
                continue
            
            # Skip if compliance reminder already sent
            if compliance_reminder_sent:
                continue
            
            # Parse welcome email timestamp
            if isinstance(welcome_email_timestamp, datetime):
                welcome_dt = welcome_email_timestamp
                # If it's naive datetime, assume UTC
                if welcome_dt.tzinfo is None:
                    welcome_dt = welcome_dt.replace(tzinfo=timezone.utc)
            else:
                # Try parsing string format
                try:
                    # Try with UTC suffix
                    welcome_dt = datetime.strptime(str(welcome_email_timestamp), "%Y-%m-%d %H:%M:%S UTC")
                    welcome_dt = welcome_dt.replace(tzinfo=timezone.utc)
                except:
                    try:
                        # Try without UTC suffix
                        welcome_dt = datetime.strptime(str(welcome_email_timestamp), "%Y-%m-%d %H:%M:%S")
                        welcome_dt = welcome_dt.replace(tzinfo=timezone.utc)
                    except:
                        print(f"⚠️ Row {row_idx}: Could not parse welcome email timestamp: {welcome_email_timestamp}")
                        continue
            
            # Check if at least 8 hours have passed since welcome email
            time_since_welcome = current_time - welcome_dt
            if time_since_welcome < timedelta(hours=8):
                continue
            
            # Get worker details
            name = row[0].value if row[0].value else "Unknown"
            email = row[1].value if len(row) > 1 and row[1].value else None
            start_date = row[13].value if len(row) > 13 and row[13].value else None  # Column N (index 13)
            
            if not email:
                print(f"⚠️ Row {row_idx}: Skipping {name} - no email address")
                continue
            
            # Calculate deadline (start date or 7 days from now as fallback)
            if start_date:
                if isinstance(start_date, datetime):
                    deadline = start_date
                else:
                    try:
                        deadline = datetime.strptime(str(start_date), "%Y-%m-%d %H:%M:%S")
                    except:
                        try:
                            deadline = datetime.strptime(str(start_date), "%Y-%m-%d")
                        except:
                            deadline = current_time + timedelta(days=7)
            else:
                deadline = current_time + timedelta(days=7)
            
            eligible_workers.append({
                'name': name,
                'email': email,
                'welcome_sent_at': welcome_dt,
                'deadline': deadline,
                'row': row_idx
            })
                
        except Exception as e:
            print(f"⚠️ Error processing row {row_idx}: {e}")
            continue
    
    wb.close()
    return eligible_workers

def _update_compliance_reminder_timestamp(row_number):
    """Update column L (index 11) with current UTC timestamp"""
    tracker_path = _resolve_tracker_path()
    wb = load_workbook(tracker_path)
    ws = wb.active
    
    timestamp = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')
    ws.cell(row=row_number, column=12, value=timestamp)  # Column L = 12
    
    wb.save(tracker_path)
    wb.close()

def send_compliance_reminder(candidate_name, deadline, recipient_email, dry_run=False):
    """
    Send compliance documents reminder email to a single candidate.
    
    Args:
        candidate_name: Name of the candidate
        deadline: datetime object for deadline
        recipient_email: Email address
        dry_run: If True, don't actually send email
    
    Returns:
        Success/failure message
    """
    try:
        # Load and populate template
        html_template = _load_html_template()
        
        # Format deadline for display
        formatted_deadline = deadline.strftime('%d-%b-%Y')
        
        html_body = html_template.replace('{Candidate_Name}', candidate_name)
        html_body = html_body.replace('{Deadline_Date}', formatted_deadline)
        
        subject = "Action Required: Sign Your Compliance Documents"
        
        if dry_run:
            print(f"[DRY RUN] Would send to: {recipient_email}")
            print(f"  Name: {candidate_name}, Deadline: {formatted_deadline}")
            return "✅ [DRY RUN] Email prepared successfully"
        
        # Send email using Gmail API
        client = GmailAPIClient()
        result = client.send_email(
            to=recipient_email,
            subject=subject,
            body=html_body
        )
        
        return result
        
    except Exception as e:
        return f"❌ Error sending compliance reminder: {str(e)}"

def batch_compliance_reminders(dry_run=None):
    """
    Process all eligible workers and send compliance document reminders.
    Updates column L after successful send.
    
    Eligibility criteria:
    - Welcome email sent (Column K has timestamp)
    - At least 8 hours since welcome email
    - Partner Domain Account triggered = "No" (Column M)
    - Compliance reminder not sent yet (Column L empty)
    
    Args:
        dry_run: If True, shows what would be sent without actually sending
    
    Returns:
        Summary of results
    """
    if dry_run is None:
        dry_run = False
        
    print("🔍 Checking for workers needing compliance document reminders...")
    
    eligible_workers = _load_eligible_workers()
    
    if not eligible_workers:
        return "ℹ️ No workers found needing compliance reminders at this time."
    
    print(f"📋 Found {len(eligible_workers)} worker(s) needing compliance reminders")
    
    results = {
        'success': [],
        'failed': []
    }
    
    for worker in eligible_workers:
        time_since_welcome = datetime.now(timezone.utc) - worker['welcome_sent_at']
        hours_since = time_since_welcome.total_seconds() / 3600
        
        print(f"\n📧 Processing: {worker['name']} ({worker['email']})")
        print(f"   Welcome sent: {worker['welcome_sent_at'].strftime('%Y-%m-%d %H:%M UTC')} ({hours_since:.1f} hours ago)")
        print(f"   Deadline: {worker['deadline'].strftime('%d-%b-%Y')}")
        
        result = send_compliance_reminder(
            candidate_name=worker['name'],
            deadline=worker['deadline'],
            recipient_email=worker['email'],
            dry_run=False
        )
        
        print(f"   {result}")
        
        if "✅" in result:
            results['success'].append({
                'row': worker['row'],
                'name': worker['name'],
                'email': worker['email']
            })
            
            # Update timestamp in Excel (only if not dry run)
            if not dry_run:
                try:
                    _update_compliance_reminder_timestamp(worker['row'])
                    print(f"   ✓ Updated timestamp in row {worker['row']}")
                except Exception as e:
                    print(f"   ⚠️ Could not update timestamp: {e}")
        else:
            results['failed'].append({
                'row': worker['row'],
                'name': worker['name'],
                'email': worker['email'],
                'error': result
            })
    
    # Summary
    summary = f"\n{'='*60}\n"
    summary += f"📊 Compliance Reminder Summary:\n"
    summary += f"   ✅ Successful: {len(results['success'])}\n"
    summary += f"   ❌ Failed: {len(results['failed'])}\n"
    
    if results['failed']:
        summary += f"\nFailed entries:\n"
        for item in results['failed']:
            summary += f"   - Row {item['row']}: {item['name']} ({item['email']})\n"
    
    print(summary)
    return summary
