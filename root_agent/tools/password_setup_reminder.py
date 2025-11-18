import os
from datetime import datetime, timezone, timedelta
from openpyxl import load_workbook
from .utils.gmail_client import GmailAPIClient
from .utils.drive_templates import load_template_from_drive
from .utils.tracker_config import get_tracker_path

def _resolve_tracker_path():
    """Return tracker path via central config."""
    return get_tracker_path()

def _load_html_template():
    """Load the Password Setup Reminder HTML template from Google Drive or local file.
    Tries Drive first, then falls back to local templates directory.
    """
    template_filename = 'Password Setup Reminder.htm'
    
    # Try loading from Google Drive first
    try:
        template_content = load_template_from_drive(template_filename, use_cache=True)
        if template_content:
            print(f"[password_setup_reminder] Loaded template from Drive: {template_filename}")
            return template_content
    except Exception as e:
        print(f"[password_setup_reminder] Drive template load failed: {e}, falling back to local")
    
    # Fallback to local templates directory
    template_dir = os.path.join(
        os.path.dirname(os.path.dirname(__file__)), 
        'templates'
    )
    template_path = os.path.join(template_dir, template_filename)
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")
    
    with open(template_path, 'r', encoding='utf-8') as f:
        print(f"[password_setup_reminder] Loaded template from local: {template_filename}")
        return f.read()

def _load_eligible_workers():
    """
    Load workers who:
    - Column R (18) = "Yes" (Password setup appointment scheduled)
    - Column S (19) has appointment time
    - Appointment time is approximately 1 hour from now
    - Column T (20) is empty (reminder not sent yet)
    
    Returns list of dicts with worker info grouped by appointment time.
    """
    tracker_path = _resolve_tracker_path()
    wb = load_workbook(tracker_path)
    ws = wb.active
    
    eligible_workers = []
    current_time = datetime.now()
    
    # Check time window: 50-70 minutes before appointment (gives 20-min buffer)
    min_reminder_time = current_time + timedelta(minutes=50)
    max_reminder_time = current_time + timedelta(minutes=70)
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        try:
            # Column R (index 17): Password setup appointment scheduled
            appointment_scheduled = row[17].value if len(row) > 17 else None
            
            # Column S (index 18): Password setup appointment time
            appointment_time = row[18].value if len(row) > 18 else None
            
            # Column T (index 19): Password setup reminder sent timestamp
            reminder_sent = row[19].value if len(row) > 19 else None
            
            # Skip if appointment not scheduled or reminder already sent
            if not appointment_scheduled or str(appointment_scheduled).strip().upper() != "YES":
                continue
            
            if reminder_sent:
                continue
            
            if not appointment_time:
                continue
            
            # Handle datetime object or string
            if isinstance(appointment_time, datetime):
                appt_dt = appointment_time
            else:
                # Try parsing string format
                try:
                    appt_dt = datetime.strptime(str(appointment_time), "%Y-%m-%d %H:%M:%S")
                except:
                    try:
                        appt_dt = datetime.strptime(str(appointment_time), "%d-%m-%Y %H:%M")
                    except:
                        print(f"⚠️ Row {row_idx}: Could not parse appointment time: {appointment_time}")
                        continue
            
            # Check if appointment is within reminder window (50-70 minutes from now)
            if min_reminder_time <= appt_dt <= max_reminder_time:
                name = row[0].value if row[0].value else "Unknown"
                email = row[1].value if len(row) > 1 and row[1].value else None
                worker_id = row[2].value if len(row) > 2 and row[2].value else "N/A"
                
                if not email:
                    print(f"⚠️ Row {row_idx}: Skipping {name} - no email address")
                    continue
                
                eligible_workers.append({
                    'name': name,
                    'email': email,
                    'worker_id': worker_id,
                    'appointment_time': appt_dt,
                    'row': row_idx
                })
                
        except Exception as e:
            print(f"⚠️ Error processing row {row_idx}: {e}")
            continue
    
    wb.close()
    return eligible_workers

def _update_reminder_timestamp(row_number):
    """Update column T (index 19) with current UTC timestamp"""
    tracker_path = _resolve_tracker_path()
    wb = load_workbook(tracker_path)
    ws = wb.active
    
    timestamp = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')
    ws.cell(row=row_number, column=20, value=timestamp)  # Column T = 20
    
    wb.save(tracker_path)
    wb.close()

def send_password_setup_reminder(candidate_name, worker_id, appointment_time, recipient_email, dry_run=False):
    """
    Send password setup reminder email to a single candidate.
    
    Args:
        candidate_name: Name of the candidate
        worker_id: Worker ID
        appointment_time: datetime object of the appointment
        recipient_email: Email address
        dry_run: If True, don't actually send email
    
    Returns:
        Success/failure message
    """
    try:
        # Load and populate template
        html_template = _load_html_template()
        
        # Format appointment time for display
        formatted_time = appointment_time.strftime('%d-%b-%Y %I:%M %p')
        
        html_body = html_template.replace('{Candidate_Name}', candidate_name)
        html_body = html_body.replace('{Worker_ID}', str(worker_id))
        html_body = html_body.replace('{Appointment_Time}', formatted_time)
        
        subject = "Reminder: Password Setup Session Starting in 1 Hour"
        
        if dry_run:
            print(f"[DRY RUN] Would send to: {recipient_email}")
            print(f"  Name: {candidate_name}, Worker ID: {worker_id}, Time: {formatted_time}")
            return "✅ [DRY RUN] Email prepared successfully"
        
        # Send email using Gmail API
        client = GmailAPIClient()
        result = client.send_email(
            to=recipient_email,
            subject=subject,
            body=html_body
        )
        
        return result
        
    except Exception as e:
        return f"❌ Error sending password setup reminder: {str(e)}"

def batch_password_setup_reminders(dry_run=False):
    """
    Process all eligible workers and send password setup reminders.
    Updates column T after successful send.
    
    Args:
        dry_run: If True, shows what would be sent without actually sending
    
    Returns:
        Summary of results
    """
    print("🔍 Checking for upcoming password setup appointments...")
    
    eligible_workers = _load_eligible_workers()
    
    if not eligible_workers:
        return "ℹ️ No password setup appointments found within the next hour."
    
    print(f"📋 Found {len(eligible_workers)} worker(s) with appointments in ~1 hour")
    
    results = {
        'success': [],
        'failed': []
    }
    
    for worker in eligible_workers:
        print(f"\n📧 Processing: {worker['name']} ({worker['email']})")
        print(f"   Appointment: {worker['appointment_time'].strftime('%d-%b-%Y %I:%M %p')}")
        
        result = send_password_setup_reminder(
            candidate_name=worker['name'],
            worker_id=worker['worker_id'],
            appointment_time=worker['appointment_time'],
            recipient_email=worker['email'],
            dry_run=dry_run
        )
        
        print(f"   {result}")
        
        if "✅" in result:
            results['success'].append({
                'row': worker['row'],
                'name': worker['name'],
                'email': worker['email']
            })
            
            # Update timestamp in Excel (only if not dry run)
            if not dry_run:
                try:
                    _update_reminder_timestamp(worker['row'])
                    print(f"   ✓ Updated timestamp in row {worker['row']}")
                except Exception as e:
                    print(f"   ⚠️ Could not update timestamp: {e}")
        else:
            results['failed'].append({
                'row': worker['row'],
                'name': worker['name'],
                'email': worker['email'],
                'error': result
            })
    
    # Summary
    summary = f"\n{'='*60}\n"
    summary += f"📊 Password Setup Reminder Summary:\n"
    summary += f"   ✅ Successful: {len(results['success'])}\n"
    summary += f"   ❌ Failed: {len(results['failed'])}\n"
    
    if results['failed']:
        summary += f"\nFailed entries:\n"
        for item in results['failed']:
            summary += f"   - Row {item['row']}: {item['name']} ({item['email']})\n"
    
    print(summary)
    return summary
